from openpyxl import load_workbook
from datetime import datetime
rut=r'BaseCrudColab.xlsx'


def borrar(ruta,identificador):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['hoja.tareas']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  descripcion=3
  estado=4
  fechaInicio=5
  fechaFinalizado=6
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      hoja.cell(row=fila,column=1).value=''
      hoja.cell(row=fila,column=nombre).value=''
      hoja.cell(row=fila,column=descripcion).value=''
      hoja.cell(row=fila,column=estado).value=''
      hoja.cell(row=fila,column=fechaInicio).value=''
      hoja.cell(row=fila,column=fechaFinalizado).value=''
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

