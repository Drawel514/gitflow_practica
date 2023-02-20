from openpyxl import load_workbook
from datetime import datetime
rut=r'BaseCrudColabProductos.xlsx'

def leer(rut:str,extraer:str): 
    Archivo_Exccel=load_workbook(rut)
    Hoja_datos=Archivo_Exccel["hoja.productos"]
    Hoja_datos=Hoja_datos["A2":"E"+str(Hoja_datos.max_row)]

    info={}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value,{"nombre":i[1].value, "categoria":i[2].value, "precio":i[3].value, "cantidad":i[4].value})

    if not(extraer=="todo"):
        info=filtrar(info, extraer)

    for i in info:
        print("****Productos****")
        print("id: "+ str(i)+"\n"+"Nombre: "+str(info[i]["nombre"])+"\n"+"Categoria: "+str(info[i]["categoria"])+"\n"+"Precio: "+str(info[i]["precio"])+"\n"+"Cantidad: "+str(info[i]["cantidad"]))
        print()

    return

def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]["categoria"]==filtro:
            aux.setdefault(i,info[i])
    return aux
def agregar(rut:int,datos:dict):
    Archivo_Exccel=load_workbook(rut)
    Hoja_datos=Archivo_Exccel["hoja.productos"]
    Hoja_datos=Hoja_datos["A2":"E"+str(Hoja_datos.max_row+1)]
    hoja=Archivo_Exccel.active

    nombre=2
    categoria=3
    precio=4
    cantidad=5
    for i in Hoja_datos:

        if not(isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=nombre).value=datos["nombre"]
            hoja.cell(row=identificador, column=categoria).value=datos["categoria"]
            hoja.cell(row=identificador, column=precio).value=datos["precio"]
            hoja.cell(row=identificador, column=cantidad).value=datos["cantidad"] 
            break
    Archivo_Exccel.save(rut)
    return

def borrar(ruta,identificador):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['hoja.productos']
  hojaDatos=hojaDatos['A2':'E'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  categoria=3
  precio=4
  cantidad=5
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      hoja.cell(row=fila,column=1).value=''
      hoja.cell(row=fila,column=nombre).value=''
      hoja.cell(row=fila,column=categoria).value=''
      hoja.cell(row=fila,column=precio).value=''
      hoja.cell(row=fila,column=cantidad).value=''
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return
