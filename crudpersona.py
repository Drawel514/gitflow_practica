from openpyxl import load_workbook
from datetime import datetime
rut=r'BaseCrudColabPersona.xlsx'

def leer(ruta:str, extraer:str):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['persona']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]

  info={}

  for i in hojaDatos:
    if isinstance(i[0].value,int):
      info.setdefault(i[0].value,{'nombre':i[1].value, 'edad':i[2].value,'telefono':i[3].value,'correo':i[4].value,'nacimiento':i[5].value})

  if not(extraer=='todo'):
    info=filtrar(info,extraer)
  for i in info:
    print('********** Tarea ***********')
    print('id:'+str(i)+'\n'+'nombre: '+str(info[i]['nombre'])+'\n'+'edad: '+str(info[i]['edad'])+'\n'+'telefono: '+str(info[i]['telefono'])+'\n'+'correo: '+str(info[i]['correo'])+'\n'+'nacimiento: '+str(info[i]['nacimiento']))
    print()
  return info

def filtrar(info:dict, filtro:str):
  aux={}
  if filtro=='mayor':

    for i in info:
      if int(info[i]['edad'])>=18:
        aux.setdefault(i,info[i])
  elif filtro=='menor':
    for i in info:
      if int(info[i]['edad'])<18:
        aux.setdefault(i,info[i])
  return aux

def actualizar(ruta:str,identificador:int,datosActualizados:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['persona']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  edad=3
  telefono=4
  correo=5
  nacimiento=6
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      for d in datosActualizados:
        if d=='nombre' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=nombre).value=datosActualizados[d]
        elif d=='edad' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=edad).value=datosActualizados[d]
        elif d=='telefono' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=telefono).value=datosActualizados[d]
        elif d=='correo' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=correo).value=datosActualizados[d]
        elif d=='nacimiento' and not(datosActualizados[d]==''):
          hoja.cell(row=fila,column=nacimiento).value=datosActualizados[d]
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una persona con ese id')
    print()
  return

def agregar(ruta:int, datos:dict):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['persona']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row+1)]
  hoja=archivoExcel.active

  nombre=2
  edad=3
  telefono=4
  correo=5
  nacimiento=6
  for i in hojaDatos:
    if not(isinstance(i[0].value,int)):
      identificador=i[0].row
      hoja.cell(row=identificador,column=1).value=identificador-1
      hoja.cell(row=identificador,column=nombre).value=datos['nombre']
      hoja.cell(row=identificador,column=edad).value=datos['edad']
      hoja.cell(row=identificador,column=telefono).value=datos['telefono']
      hoja.cell(row=identificador,column=correo).value=datos['correo']
      hoja.cell(row=identificador,column=nacimiento).value=datos['nacimiento']
      break
  archivoExcel.save(ruta)
  return

def borrar(ruta,identificador):
  archivoExcel=load_workbook(ruta)

  hojaDatos=archivoExcel['persona']
  hojaDatos=hojaDatos['A2':'F'+str(hojaDatos.max_row)]
  hoja=archivoExcel.active

  nombre=2
  edad=3
  telefono=4
  correo=5
  nacimiento=6
  encontro=False
  for i in hojaDatos:
    if i[0].value==identificador:
      fila=i[0].row
      encontro=True
      hoja.cell(row=fila,column=1).value=''
      hoja.cell(row=fila,column=nombre).value=''
      hoja.cell(row=fila,column=edad).value=''
      hoja.cell(row=fila,column=telefono).value=''
      hoja.cell(row=fila,column=correo).value=''
      hoja.cell(row=fila,column=nacimiento).value=''
  archivoExcel.save(ruta)
  if encontro == False:
    print('Error: no existe una tarea con ese id')
    print()
  return

while True:
  print('Indique la accion que desea realizar: \nConsultar: 1\nActualizar: 2\nRegistrar: 3\nBorrar: 4')
  accion =int(input('Escriba la opcion: '))
  print('*****************')
  if accion<1 or accion>4:
    print('Comando invalido, por favor eliga una opcion valida')
  elif accion==1:
    opcConsulta=''
    print('Indique la opcion que desee:\nTodas las personas: 1\nMayores o iguales 18: 2\nMenores a 18: 3')
    opcConsulta=input('Escriba la persona que see consultar: ')
    if opcConsulta=='1':
      print('\n\n** Consultado todas las personas **')
      leer(rut,'todo')
    elif opcConsulta=='2':
      print('\n\n** Consultado todas las personas **')
      leer(rut,'mayor')
    elif opcConsulta=='3':
      print('\n\n** Consultado todas las personas **')
      leer(rut,'menor')
  elif accion==2:
    datosActualizados={'nombre':'','edad':'','telefono':'','correo':'','nacimiento':''}
    print('** Actualizar persona **\n')
    idActualizar=int(input('Indique el ID de la persona que desea actualizar: '))
    print('\n** Nuevo nombre **\n** Nota: si no desea actualizar el nombre solo oprima ENTER **')
    datosActualizados['nombre']=input('Indique el nuevo nombre de la persona: ')
    print('\n** Nueva edad **\n** Nota: si no desea actualizar la edad solo oprima ENTER **')
    datosActualizados['edad']=input('Indique la nueva edad de la persona: ')
    print('\n** Nuevo telefono **')
    datosActualizados['telefono']=input('Indique el nuevo telefono de la persona: ')
    print('\n** Nuevo correo **')
    datosActualizados['correo']=input('Indique el nuevo correo de la persona: ')
    print('\n** Nuevo nacimiento **')
    datosActualizados['nacimiento']=input('Indique el nuevo nacimiento de la persona: ')
    actualizar(rut,idActualizar, datosActualizados)
    print()
  elif accion==3:
    datosActualizados={'nombre':'','edad':'','telefono':'','correo':'','nacimiento':''}
    print('** Crear nueva persona **\n')
    print('** nombre **\n')
    datosActualizados['nombre']=input('Indique el nombre de la persona: ')
    print('\n** edad **')
    datosActualizados['edad']=input('Indique la edad de la persona: ')
    print()
    print('\n** telefono **')
    datosActualizados['telefono']=input('Indique el telefono de la persona: ')
    print()
    print('\n** correo **')
    datosActualizados['correo']=input('Indique el correo de la persona: ')
    print()
    print('\n** nacimiento **')
    datosActualizados['nacimiento']=input('Indique el nacimiento de la persona: ')
    agregar(rut,datosActualizados)

  elif accion==4:
    print('\n** Eliminar persona **')
    iden=int(input('Indique el ID de la persona que desea eliminar: '))
    borrar(rut,iden)
    print('*************')